# coding: utf-8

"""
Programme d'identification de doublons entre plusieurs fichiers
Cas envisagé : plusieurs fichiers au même format (tabulé, colonnes aux mêmes positions)
contenant une liste d'exemplaires en colonne X.
Le programme ouvre chacun des fichiers et récupère l'ensemble des codes-barres, qu'il associe
dans un dictionnaire à ce fichier

Puis il ouvre le fichier suivant et enrichit le même dictionnaire.
S'il rencontre le même code-barre, il va enrichir l'entrée déjà existante du dictionnaire

Les fichiers en entrée ont des en-têtes de colonne

Le dictionnaire ressemble à :
dic_ids = {
            "14727841318": ("fichier1.txt"),
            "14725418718": ("fichier1.txt"),
            "71561516510": ("fichier1.txt", "fichier2.txt"),
            "12217785424": ("fichier2.txt")
            } 

Ensuite, le programme extrait du dictionnaire les valeurs multiples

"""

import string
import os 
from collections import defaultdict
import csv
from openpyxl import Workbook, load_workbook

dir_path = os.path.dirname(os.path.realpath(__file__))

#dic_ids associe chaque identifiant à la liste des fichiers où on le trouve
dic_ids = defaultdict(set)

# dic_ids2metas enregistre, pour un ID dans un fichier, ses métadonnées initiales
# ce dictionnaire n'est alimenté que si on choisit de conserver les métadonnées
dic_id2metas = defaultdict(list)

# dic_empty_ids_counter compte les ID vides dans chaque fichier
dic_empty_ids_counter = defaultdict(int)

def representsInt(s):
    """
    Vérifie si s est un nombre
    """
    test = True
    try:
        int(s)
    except ValueError:
        test = False
    return test


def select_column(select_input, headers):
    """Vérifie si la valeur du 2e input est un chiffre (n° de colonne)
    ou non (en-tête de colonne)"""
    if (select_input == ""):
        return 0
    if (representsInt(select_input)):
        return int(select_input)-1
    else:
        return headers.index(select_input)


def file2dic(filename, select_col, recup_meta, unsignificant_values, unsignificant_values_col):
    if ("xls" in filename):
        # ouverture du fichier Excel
        excel_file2dict(filename, select_col, recup_meta, unsignificant_values, unsignificant_values_col)
    else:
        csv_file2dict(filename, select_col, recup_meta, unsignificant_values, unsignificant_values_col)


def select_column_xls(select_input, headers):
    col = 0
    if (representsInt(select_input)):
        col = int(select_input)-1
    else:
        i = 0
        for cell in headers:
            if (cell.value == select_input):
                col = i
            i += 1
    # Conversion de l'indice de colonne en lettre de colonne
    col = string.ascii_uppercase[col]
    return col


def row_validator(identifier, table, row, 
                  unsignificant_values, unsignificant_values_col_id,
                  input_filetype):
    """Si la ligne contient (dans une colonne à indiquer en paramètre)
    une certaine valeur, alors l'identifiant est considéré comme vide
    et la ligne n'est pas traitée"""
    if ("xlsx" in input_filetype):
        unsignif_cell_name = f"{unsignificant_values_col_id}{str(row)}"
        unsignif_cell_name_val = table[unsignif_cell_name].value
    else:
        unsignif_cell_name_val = row[unsignificant_values_col_id]
    for el in unsignificant_values:
        if unsignif_cell_name_val == el:
            identifier = ""
    return identifier


def excel_file2dict(input_filename, select_col, recup_meta, unsignificant_values, unsignificant_values_col):
    xlsfile = load_workbook(filename=input_filename)
    xls_table_name = xlsfile.sheetnames[0]
    xls_table = xlsfile[xls_table_name]
    column_id = select_column_xls(select_col, list(xls_table.rows)[0])
    unsignificant_values_col_id = select_column_xls(unsignificant_values_col, 
                                                    list(xls_table.rows)[0])
    for row in range(2, xls_table.max_row + 1):
        cell_name = f"{column_id}{str(row)}"
        identifier = xls_table[cell_name].value
        print("open file", input_filename, identifier)
        identifier = str(row_validator(identifier, xls_table, row, 
                                   unsignificant_values, unsignificant_values_col_id,
                                   "xlsx"))
        if (identifier):
            dic_ids[identifier].add(filename)
            if (recup_meta == "o"):
                key = identifier + input_filename
                for col in range(0, xls_table.max_column):
                    cell = string.ascii_uppercase[col]+str(row)
                    dic_id2metas[key].append(xls_table[cell].value)
        else:
            # Si la zone Identifiant est vide, 
            # on l'ajoute au compteur des ID vides
            dic_empty_ids_counter[input_filename] += 1


def csv_file2dict(input_filename, select_col, recup_meta, unsignificant_values, unsignificant_values_col):
    with open(input_filename, encoding="utf-8") as csvfile:
            content  = csv.reader(csvfile, delimiter='\t')
            headers = next(content)
            column_id = select_column(select_col, headers)
            unsignificant_values_col_id = select_column(unsignificant_values_col_col, 
                                                        headers)

            for row in content:
                identifier = row[column_id]
                print("open file", input_filename, identifier)
                identifier = str(row_validator(identifier, content, row, 
                                           unsignificant_values, unsignificant_values_col_id,
                                           "csv"))
                if (identifier):
                    dic_ids[identifier].add(input_filename)
                    if (recup_meta == "o"):
                        dic_id2metas[identifier+input_filename] = row
                else:
                    # Si la zone Identifiant est vide, 
                    # on l'ajoute au compteur des ID vides
                    dic_empty_ids_counter[input_filename] += 1

    

def create_csv_report(output_filename, output_filetype, recup_meta):
    report = open(output_filename, "w", encoding="utf-8")
    report.write("\t".join["Identifiant doublon", "fichiers concernés"] + "\n")
    return report


def create_xls_report(output_filename, output_filetype, recup_meta):
    report = Workbook(write_only=False)
    report.guess_type = False
    report_table = report.create_sheet(title="liste_doublons")
    if ("Sheet" in report.get_sheet_names()):
        sh1 = report.get_sheet_by_name("Sheet")
        report.remove_sheet(sh1)
    report_table["A1"] = "Identifiant doublon"
    report_table["B1"] =  "fichiers concernés"
    report.save(os.path.join(dir_path, output_filename))
    return report, report_table
        


def line2report(line, report, output_filetype, i):
    if (output_filetype == "csv"):
        report.write("\t".join(line) + "\n")
    else:
        j = 0
        for el in line:
            cell_adress = string.ascii_uppercase[j]+str(i)
            report[cell_adress] = str(line[j])
            j += 1


def dict_entry2report(report, entry, entry_value, i, output_filetype, recup_meta):
    for source in list(dic_ids[entry]):
        i += 1
        line = [entry, source]
        if (recup_meta == "o"):
            key = entry + source
            meta = dic_id2metas[key]
            line.extend(meta)
        print(line)
        line2report(line, report, output_filetype, i)
    return i
    

def report_empty_ids():
    if dic_empty_ids_counter:
        print("\n\n", "-"*15,"\nDécompte des lignes avec identifiant vide, par fichier\n\
(sont comptabilisées les lignes ignorées parce que contenant\n\
une valeur spécifique indiquée par l'utilisateur\n",
              "-"*15)
        print("Nom du fichier\tNombre de lignes vides")
        for filename in dic_empty_ids_counter:
            print(filename, "\t", dic_empty_ids_counter[filename])


if __name__ =="__main__":
    filelist = input("\nNom des fichiers à comparer (séparés par des ';') : ")
    select_col = input("\nNom ou numéro de colonne (numérotation commençant à 1) \
servant d 'identifiant (par défaut : 1ère colonne) : ")
    unsignificant_values = input('\nIgnorer certaines valeurs ?\n\
(permet de préciser des lignes à ne pas prendr en compte - séparateur ";" : ').split(";")
    unsignificant_values_col = input("\nColonne où se trouvent les valeurs à ne pas prendre\n\
en compte (numérotation commençant à 1): ")
    recup_meta = input("\nRécupérer toutes les métadonnées (O/N) ? ").lower()
    if (recup_meta == ""):
        recup_meta = "o"
    output_filename = input("Nom du rapport de doublons : ")
    output_filetype = "csv"
    for filename in filelist.split(";"):
        file2dic(filename, select_col, recup_meta, unsignificant_values, unsignificant_values_col)

    if (".xls" in output_filename):
        output_filetype = "xlsx"
        report, sheet = create_xls_report(output_filename, output_filetype, recup_meta)
    else:
        report = create_csv_report(output_filename, output_filetype, recup_meta)

    print("\n\n", "-"*15, "\nRecherche des doublons\n", "-"*15)
    i = 1
    if (".xls" in output_filename):
        for entry in dic_ids:
            if len(dic_ids[entry]) > 1:
                i = dict_entry2report(sheet, entry, dic_ids[entry], i, output_filetype, recup_meta)
                
        report.save(os.path.join(dir_path, output_filename))
    else:
        for entry in dic_ids:
            if len(dic_ids[entry]) > 1:
                i = dict_entry2report(report, entry, dic_ids[entry], i, output_filetype, recup_meta)
    report_empty_ids()